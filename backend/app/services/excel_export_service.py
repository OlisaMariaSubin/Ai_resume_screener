"""Recruiter-facing Excel export (Section 1). Builds a multi-sheet .xlsx workbook from
data screening_service/trend_analysis have already computed - never invents a value,
and every optional field is read defensively so a resume missing a field (degree,
branch, experience...) never crashes generation.
"""
import io

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

SCORE_BUCKETS = [
    ("Strong Match (80-100)", 80, 100),
    ("Moderate Match (60-79)", 60, 79),
    ("Weak Match (0-59)", 0, 59),
]


def _screening_decision(overall: float | None) -> str:
    if overall is None:
        return ""
    for label, low, high in SCORE_BUCKETS:
        if low <= overall <= high:
            return label
    return ""


def _join(values) -> str:
    return ", ".join(values) if values else ""


def _results_rows(results: list[dict], resume_lookup: dict[str, dict]) -> list[dict]:
    """Sheet 1 - only ELIGIBLE candidates who were actually screened/scored. Ineligible
    candidates never appear here (Section 1 'Important').

    resume_lookup: candidate_id -> resume structured_data dict (may be missing a key
    entirely if the resume row is gone - handled with .get() throughout).
    """
    rows = []
    for r in results:
        if r.get("status") != "scored":
            continue
        score = r.get("score") or {}
        skills = r.get("skills") or {}
        structured = resume_lookup.get(r.get("candidate_id"), {})
        rows.append(
            {
                "Candidate Name": r.get("candidate_name") or r.get("filename") or "",
                "Resume/File Name": r.get("filename", ""),
                "Eligibility Status": r.get("eligibility_status") or "ELIGIBLE",
                "Eligibility Reason": r.get("eligibility_reason") or "",
                "Degree": r.get("extracted_degree", ""),
                "Branch/Specialization": r.get("extracted_branch", ""),
                "Experience": _join(structured.get("experience")),
                "Skills": _join(structured.get("skills")),
                "Match Score": score.get("overall"),
                "Screening Decision": _screening_decision(score.get("overall")),
                "Missing Skills": _join(skills.get("missing")),
                "Matching Skills": _join(skills.get("matched")),
                "Explanation": r.get("explanation") or "",
                "Overqualified": "Yes" if r.get("overqualified") else "No",
                "Overqualification Reason": r.get("overqualification_reason") or "",
                "Ranking": r.get("ranking"),
                "Skill Match %": score.get("skill_match_pct"),
                "TF-IDF Similarity": score.get("tfidf_similarity"),
                "Embedding Similarity": score.get("embedding_similarity"),
            }
        )
    return rows


def _rejected_rows(results: list[dict], job) -> list[dict]:
    """Sheet 2 - candidates filtered out during eligibility pre-screening."""
    required_degree = ""
    if job.eligibility_config and job.eligibility_config.get("requirement_type") not in (None, "none"):
        required_degree = " / ".join(job.eligibility_config.get("allowed_degree_names") or []) or (
            job.eligibility_config.get("minimum_degree_level") or ""
        )

    rows = []
    for r in results:
        if r.get("status") != "ineligible":
            continue
        rows.append(
            {
                "Candidate": r.get("candidate_name") or r.get("filename") or "",
                "Resume": r.get("filename", ""),
                "Extracted Degree": r.get("extracted_degree", ""),
                "Required Degree": r.get("required_degree_text") or required_degree,
                "Eligibility Status": r.get("eligibility_status") or "INELIGIBLE",
                "Rejection Reason": r.get("eligibility_reason") or "",
                "Overqualified": "Yes" if r.get("overqualified") else "No",
                "Overqualification Reason": r.get("overqualification_reason") or "",
            }
        )
    return rows


def _trend_rows(trends: list[dict]) -> list[dict]:
    return [
        {
            "Skill": t.get("skill", ""),
            "Candidate Count": t.get("candidates_with_skill", 0),
            "Percentage of Pool": t.get("pct_of_pool", 0.0),
            "Rarity/Commonality": t.get("rarity_category", ""),
            "JD Must-Have": "Yes" if t.get("is_jd_must_have") else "No",
            "JD Nice-to-Have": "Yes" if t.get("is_jd_nice_to_have") else "No",
        }
        for t in trends
    ]


def _insight_rows(insights: dict) -> list[dict]:
    rows = [
        {"Category": "Oversupplied skill (Very Common)", "Detail": skill, "Insight": ""}
        for skill in insights.get("oversupplied_skills", [])
    ]
    rows += [
        {
            "Category": f"Rare skill ({r['rarity_category']})",
            "Detail": f"{r['skill']} ({r['pct_of_pool']}% of pool)",
            "Insight": "",
        }
        for r in insights.get("rare_skills", [])
    ]
    rows += [
        {"Category": "JD Requirement Gap", "Detail": gap["skill"], "Insight": gap["insight"]}
        for gap in insights.get("jd_requirement_gaps", [])
    ]
    return rows


def _autosize_and_style(ws) -> None:
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in column_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)


def _write_sheet(writer, sheet_name: str, rows: list[dict], columns: list[str], empty_note: str) -> None:
    df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    _autosize_and_style(ws)
    if not rows:
        ws.cell(row=2, column=1, value=empty_note)


def build_screening_workbook(
    job, results: list[dict], trends: list[dict], insights: dict, resume_lookup: dict[str, dict] | None = None
) -> io.BytesIO:
    """Returns an in-memory .xlsx workbook (BytesIO, positioned at 0) with:
    Sheet 1 Screening Results - eligible/scored candidates only.
    Sheet 2 Eligibility Rejected - candidates filtered out pre-screening.
    Sheet 3 Trends - skill rarity/commonality across the eligible pool.
    Sheet 4 Recruiter Insights - JD requirement gaps and oversupplied/rare summaries.
    Never raises on missing optional fields - every field is read with .get()/defaults.
    """
    results_rows = _results_rows(results, resume_lookup or {})
    rejected_rows = _rejected_rows(results, job)
    trend_rows = _trend_rows(trends)
    insight_rows = _insight_rows(insights)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        _write_sheet(
            writer,
            "Screening Results",
            results_rows,
            [
                "Candidate Name",
                "Resume/File Name",
                "Eligibility Status",
                "Eligibility Reason",
                "Degree",
                "Branch/Specialization",
                "Experience",
                "Skills",
                "Match Score",
                "Screening Decision",
                "Missing Skills",
                "Matching Skills",
                "Explanation",
                "Overqualified",
                "Overqualification Reason",
                "Ranking",
                "Skill Match %",
                "TF-IDF Similarity",
                "Embedding Similarity",
            ],
            "No eligible candidates were screened for this job yet.",
        )
        _write_sheet(
            writer,
            "Eligibility Rejected",
            rejected_rows,
            [
                "Candidate",
                "Resume",
                "Extracted Degree",
                "Required Degree",
                "Eligibility Status",
                "Rejection Reason",
                "Overqualified",
                "Overqualification Reason",
            ],
            "No candidates were rejected during eligibility pre-screening.",
        )
        _write_sheet(
            writer,
            "Trends",
            trend_rows,
            ["Skill", "Candidate Count", "Percentage of Pool", "Rarity/Commonality", "JD Must-Have", "JD Nice-to-Have"],
            "No skill trend data available (no eligible candidates screened yet).",
        )
        _write_sheet(
            writer,
            "Recruiter Insights",
            insight_rows,
            ["Category", "Detail", "Insight"],
            "No recruiter insights available yet.",
        )

    buffer.seek(0)
    return buffer
