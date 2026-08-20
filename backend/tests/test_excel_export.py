import io

import openpyxl

from app.models.job import Job
from app.services import excel_export_service
from tests.conftest import make_docx_bytes

EXPECTED_SHEETS = ["Screening Results", "Eligibility Rejected", "Trends", "Recruiter Insights"]


def _job() -> Job:
    job = Job()
    job.title = "Backend Engineer"
    job.eligibility_config = {
        "requirement_type": "mandatory",
        "minimum_degree_level": "BACHELORS",
        "allowed_degree_names": ["B.Tech"],
        "allow_equivalent": False,
        "allow_higher_degree": False,
        "source_text": "",
    }
    return job


SCORED_RESULT = {
    "candidate_id": "c1",
    "candidate_name": "Alex Candidate",
    "filename": "alex.docx",
    "status": "scored",
    "failure_reason": None,
    "score": {
        "overall": 82.5,
        "skill_match_pct": 90.0,
        "tfidf_similarity": 0.7,
        "embedding_similarity": None,
        "weights_used": {"skill_match": 50, "text_similarity": 30, "experience": 10, "education": 10},
    },
    "skills": {"matched": ["Python", "SQL"], "missing": ["Docker"], "missing_must_have": [], "missing_nice_to_have": ["Docker"]},
    "ranking": 1,
    "explanation": None,
    "eligibility_status": "ELIGIBLE",
    "eligibility_reason": None,
    "overqualified": False,
    "overqualification_reason": None,
    "extracted_degree": "B.Tech",
    "extracted_branch": "Computer Science",
    "required_degree_text": "B.Tech",
}

INELIGIBLE_RESULT = {
    "candidate_id": "c2",
    "candidate_name": "Sam Candidate",
    "filename": "sam.docx",
    "status": "ineligible",
    "failure_reason": None,
    "score": None,
    "skills": None,
    "ranking": None,
    "eligibility_status": "INELIGIBLE",
    "eligibility_reason": "Required degree (B.Tech) not satisfied; candidate's highest listed degree is M.Tech.",
    "overqualified": True,
    "overqualification_reason": "Candidate's highest listed degree is M.Tech (Master's), above the required level.",
    "extracted_degree": "M.Tech",
    "extracted_branch": "",
    "required_degree_text": "B.Tech",
}


def _load(buffer: io.BytesIO):
    buffer.seek(0)
    return openpyxl.load_workbook(buffer)


def test_excel_export_has_expected_sheets():
    wb = _load(
        excel_export_service.build_screening_workbook(
            _job(), [SCORED_RESULT, INELIGIBLE_RESULT], trends=[], insights={}
        )
    )
    assert wb.sheetnames == EXPECTED_SHEETS


def test_excel_export_only_includes_eligible_scored_candidates_in_results_sheet():
    wb = _load(
        excel_export_service.build_screening_workbook(
            _job(), [SCORED_RESULT, INELIGIBLE_RESULT], trends=[], insights={}
        )
    )
    ws = wb["Screening Results"]
    names = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]
    assert names == ["Alex Candidate"]
    assert "Sam Candidate" not in names


def test_excel_export_rejected_sheet_has_ineligible_candidate():
    wb = _load(
        excel_export_service.build_screening_workbook(
            _job(), [SCORED_RESULT, INELIGIBLE_RESULT], trends=[], insights={}
        )
    )
    ws = wb["Eligibility Rejected"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    assert rows[0][0] == "Sam Candidate"
    assert rows[0][2] == "M.Tech"  # Extracted Degree
    assert rows[0][4] == "INELIGIBLE"  # Eligibility Status


def test_excel_export_all_ineligible_produces_empty_results_sheet_without_crashing():
    """Test 6: export must succeed even when every candidate is ineligible."""
    wb = _load(excel_export_service.build_screening_workbook(_job(), [INELIGIBLE_RESULT], trends=[], insights={}))
    ws = wb["Screening Results"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert data_rows[0][0] is not None  # the "no eligible candidates" note
    assert wb["Eligibility Rejected"]["A2"].value == "Sam Candidate"


def test_excel_export_missing_optional_fields_does_not_crash():
    """Test 9: a scored result missing optional fields (no explanation, no degree/
    branch, embedding_similarity None) must not raise during export."""
    sparse_result = dict(SCORED_RESULT)
    sparse_result["extracted_degree"] = ""
    sparse_result["extracted_branch"] = ""
    sparse_result["explanation"] = None
    sparse_result["score"] = dict(SCORED_RESULT["score"])
    sparse_result["score"]["embedding_similarity"] = None

    wb = _load(excel_export_service.build_screening_workbook(_job(), [sparse_result], trends=[], insights={}))
    ws = wb["Screening Results"]
    row = next(ws.iter_rows(min_row=2, values_only=True))
    assert row[0] == "Alex Candidate"


def test_excel_export_empty_results_does_not_crash():
    wb = _load(excel_export_service.build_screening_workbook(_job(), [], trends=[], insights={}))
    assert wb["Screening Results"]["A2"].value
    assert wb["Trends"]["A2"].value


def test_excel_export_trends_sheet_contains_rarity():
    trends = [{"skill": "Python", "candidates_with_skill": 8, "pct_of_pool": 80.0, "rarity_category": "Very Common", "is_jd_must_have": True, "is_jd_nice_to_have": False}]
    insights = {"oversupplied_skills": ["Python"], "rare_skills": [], "jd_requirement_gaps": []}
    wb = _load(excel_export_service.build_screening_workbook(_job(), [SCORED_RESULT], trends, insights))
    ws = wb["Trends"]
    row = next(ws.iter_rows(min_row=2, values_only=True))
    assert row[0] == "Python"
    assert row[3] == "Very Common"


# ---------------------------------------------------------------------------
# API-level: full endpoint (Test 5)
# ---------------------------------------------------------------------------


def test_api_export_excel_endpoint(client):
    job_resp = client.post(
        "/api/jobs", json={"title": "Data Engineer", "description": "Required:\nPython\nSQL\n"}
    )
    job_id = job_resp.json()["job_id"]

    docx_bytes = make_docx_bytes(["Pat Candidate"], table_rows=[["Skills"], ["Python, SQL"]])
    resume_id = client.post(
        "/api/resumes",
        files={"file": ("pat.docx", docx_bytes, "application/vnd.openxmlformats-officedocument.wordprocessingml.document")},
    ).json()["resume_id"]
    client.post("/api/screen", json={"job_id": job_id, "resume_id": resume_id})

    response = client.get(f"/api/screen/{job_id}/export/excel")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in response.headers["content-disposition"]
    assert ".xlsx" in response.headers["content-disposition"]

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == EXPECTED_SHEETS
    ws = wb["Screening Results"]
    assert ws["A2"].value == "Pat Candidate"


def test_api_export_excel_empty_job_does_not_crash(client):
    job_resp = client.post("/api/jobs", json={"title": "Empty Job", "description": "Required:\nPython\n"})
    job_id = job_resp.json()["job_id"]

    response = client.get(f"/api/screen/{job_id}/export/excel")
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb["Screening Results"]["A2"].value
